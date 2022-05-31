from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active

# ws['A1'] = "babangida2323"
# ws['A2'] = "babangida223"
# ws['A3'] = "babangida22322"
# ws['A4'] = "babangida223222333"
# ws.append([1, 2, 3])
# ws['A12'] = datetime.datetime.now()
testdata1 = [['data1', '---'],['data2', '-----'],['data1', '---'],['data2', '-----'],['data1', '---'],['data2', '-----']]

# for data in testdata1:
#     ws.append(data)

for i in range(1, 1000):
    for j in range(1, 999):
        ws.cell(row=i, column=j).value = i+j

wb.save("papata11.xlsx")

