from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
fake_data = Faker(['it_IT', 'en_US', 'bg_BG'])

for i in range(1, 101):
    for j in range(1, 4):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email()
        ws.cell(row=i, column=3).value = fake_data.address()
wb.save("testData.xlsx")
